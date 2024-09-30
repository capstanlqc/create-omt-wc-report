import time
import os
import sys
import argparse
from glob import glob
import json
from rich import print
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import mimetypes

def get_config(config_fpath):

    # provide config with arg -c ?
    with open(config_fpath) as f:
        config = json.load(f)

    return config["project_filter"], config["file_filter"]
    

def is_json_file(file_path):
    if not os.path.isfile(file_path):
        return False
    
    mime_type, _ = mimetypes.guess_type(file_path)
    if mime_type != 'application/json':
        return False
    
    try:
        with open(file_path, 'r') as file:
            json.load(file)  # Try to parse the JSON
        return True  # If successful, it's a valid JSON file
    except (json.JSONDecodeError, IOError):
        return False  # If there was an error, it's not valid JSON


def adjust_cols(excel_fpath):

    # Load the workbook and select the active worksheet
    workbook = load_workbook(excel_fpath)
    # worksheet = workbook.active

    # check and remove empty sheets
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # align left first column
        for row in worksheet.iter_rows(min_col=1, max_col=1):  # only the first column
            for cell in row:
                cell.alignment = Alignment(horizontal='left')

        # adjust column widths to fit the contents
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # get the column letter (A, B, C, etc.)

            if column[0].value:
                max_length = len(str(column[0].value))  # Header length

            for cell in column:
                try:
                    # Update max length if the cell value is longer
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)  # Adding some extra space
            worksheet.column_dimensions[column_letter].width = adjusted_width        
                
        # check if the sheet is empty (no rows or columns with data)
        if worksheet.max_row == 1 and worksheet.max_column == 1:
            workbook.remove(worksheet)
            print(f"Removed empty sheet: {sheet}")

    # ave the workbook with adjusted column widths
    workbook.save(excel_fpath)


def extract_wc_data(data):
    return {
        "project": 
            {
                "total": data["total"]["words"],
                "remaining": data["remaining"]["words"],
                "unique": data["unique"]["words"],
                "unique-remaining": data["unique-remaining"]["words"]
            }
        }


def add_to_workbook(data, target_lang, wb_fpath):
    if not os.path.exists(wb_fpath):
        workbook = openpyxl.Workbook()
        workbook.save(wb_fpath) # empty wb

    df = pd.DataFrame.from_dict(data, orient='index')

    # load the existing workbook and append to it without altering other sheets
    with pd.ExcelWriter(wb_fpath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=target_lang, index=True)


if __name__ != "__main__":
    sys.exit()

parser = argparse.ArgumentParser(
    prog="create-omt-wc-report", description="Create wordcount report for one or more OmegaT projects"
)
parser.add_argument("-p", "--path", help="Path to project or to folder containing projects")
parser.add_argument("-r", "--report", help="Path to the spreadsheet to save wordcounts")
parser.add_argument("-v", "--version", action="version", version="%(prog)s 0.1")
args = parser.parse_args()

if not args.path or not args.report:
    parser.print_help(sys.stderr)
    sys.stderr.write(f"\nERROR: Path '{args.path}' is not valid!\n")
    sys.exit(1)

folder_dpath = args.path
report_fpath = args.report
project_filter, file_filter = get_config("config.json")

# todo: check if folder is omegat project or not
# let's assume it's not

omtprj_fpaths = [
    path for path in glob(f"{folder_dpath}/**/omegat.project") # , recursive=True)
    if any(s in path for s in project_filter)
]

print(f"Filtered: {omtprj_fpaths}")

for omtprj_fpath in omtprj_fpaths:
    data = {}
    omtprj_dpath = os.path.dirname(omtprj_fpath)
    
    print(omtprj_dpath)
    stats_fpath = os.path.join(omtprj_dpath, "omegat", "project_stats.json")
    if not os.path.isfile(stats_fpath):
        continue # or run console-translate omegat on the project

    if not is_json_file(stats_fpath):
        continue

    with open(stats_fpath) as f:
        data = json.load(f)

    # if config["get_stats_per_file"]:
    #     print("process files...")

    if len(file_filter) > 0:
        
        subset_stats = {}
        for subset in file_filter:
            files = [f for f in data["files"] if subset in f["filename"]]
            subset_stats[subset] = {}
            subset_stats[subset].update({"total": sum([f["total"]["words"] for f in files])})
            subset_stats[subset].update({"remaining": sum([f["remaining"]["words"] for f in files])})
            subset_stats[subset].update({"unique": sum([f["unique"]["words"] for f in files])})
            subset_stats[subset].update({"unique-remaining": sum([f["unique-remaining"]["words"] for f in files])})


    wc_data = extract_wc_data(data)
    wc_data.update(subset_stats)

    print(data["project"]["target-language"])
    add_to_workbook(wc_data, data["project"]["target-language"], report_fpath)


adjust_cols(report_fpath)

# todo:
# unzip omegat project packages? 
# config option to output one single report for all or one single one for each
# process files if config["get_stats_per_file"]: true
